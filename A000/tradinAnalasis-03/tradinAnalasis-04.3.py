import time
import datetime
from matplotlib import pyplot as plt
from binance import Client
import pandas
from tradingview_ta.main import Interval, TA_Handler

test = True

class excel:
    def excelprint(what, col, row):
        import openpyxl
        wb_obj = openpyxl.load_workbook("data/balanceData.xlsx")
        sheet_obj = wb_obj.active
        sheet_obj.cell(row=row, column=col, value=what)
        wb_obj.save("data/balanceData.xlsx")
        return what
    def excelread(col, row):
        import openpyxl
        wb_obj = openpyxl.load_workbook("data/balanceData.xlsx")
        sheet_obj = wb_obj.active
        return(sheet_obj.cell(row=row, column=col).value)

currency = []
decimal = []
for x in range (2, 24):
    currency.append(excel.excelread(1, x))
    decimal.append(excel.excelread(2, x))

exchange = 'USDT'
dateback = datetime.datetime(2013,1,1)
datebackshort = datetime.datetime(2021,6,5)
plt.figure(figsize=(15, 3))

def setupNote(key):
    import json
    with open('data/setup.txt', 'r') as f:
        a = json.loads(f.read())[key]
        return a
class openordertext():
    def openorder():
        import json
        with open('data/openorder.txt', 'r') as f:
            a = json.loads(f.read())
            return a
    def writeOpenOrder(myJSONobject):
        import json
        with open("data/openorder.txt", "w") as output:
            json.dump(myJSONobject, output)
    def writeandapend(numlistcurrency):
        lis = openordertext.openorder()
        lis.append(numlistcurrency)
        openordertext.writeOpenOrder(lis)
    def removelist(numlistcurrency):
        lis = openordertext.openorder()
        lis.remove(numlistcurrency)
        openordertext.writeOpenOrder(lis)
class shortopenorder():
    def openorder():
        import json
        with open('data/shortopenorder.txt', 'r') as f:
            a = json.loads(f.read())
            return a
    def writeOpenOrder(myJSONobject):
        import json
        with open("data/shortopenorder.txt", "w") as output:
            json.dump(myJSONobject, output)
    def writeandapend(numlistcurrency):
        lis = shortopenorder.openorder()
        lis.append(numlistcurrency)
        shortopenorder.writeOpenOrder(lis)
    def removelist(numlistcurrency):
        lis = shortopenorder.openorder()
        lis.remove(numlistcurrency)
        shortopenorder.writeOpenOrder(lis)

client = Client(setupNote('key'), setupNote('secret'))

class data():
    def __init__(self, numlistcurrency):
        coinshort = TA_Handler(symbol=(str(currency[numlistcurrency]) + str(exchange)), screener="crypto", exchange="BINANCE", interval=Interval.INTERVAL_1_DAY)
        self.summaryshort = coinshort.get_analysis().summary
        self.notsummanry = coinshort.get_analysis().summary
        self.RECOMMENDATION = self.summaryshort['RECOMMENDATION']
        self.balanceCoin =  client.get_asset_balance(currency[numlistcurrency])['free']
        self.balanceUSDT =  client.get_asset_balance(exchange)['free']
        try: self.price = ((client.get_historical_klines(currency[numlistcurrency] + exchange, Client.KLINE_INTERVAL_1MINUTE, "1 minute ago UTC"))[0])[1]
        except: self.price = 0

def excelData():
    from console_progressbar import ProgressBar
    pb = ProgressBar(total=len(currency)-1,prefix='LoadData', suffix='', decimals=3, length=50, fill='X', zfill='-')
    for x in range(len(currency)):
        excel.excelprint(float(data(x).balanceCoin), 3, x+2)
        excel.excelprint(float(data(x).price), 4, x+2)
        pb.print_progress_bar(x)
    excel.excelprint(round(float(data(1).balanceUSDT)), 5, len(currency)+3)
    print("")

def lastword(string):
    lis = list(string.split(":"))
    length = len(lis)
    return lis[length-1]
def printError():
    import traceback
    print(traceback.format_exc())
    linesend(datetimestrft() + lastword(traceback.format_exc()))
    time.sleep(10)
def datetimestrft():
    return datetime.datetime.now().strftime("%d %B %Y // %H:%M:%S //")
def linesend(message):
    import numpy as np
    import requests, urllib.parse
    import io
    from PIL import Image
    import os
    token = setupNote('line')
    url = 'https://notify-api.line.me/api/notify'
    HEADERS = {'Authorization': 'Bearer ' + token}
    msg = message
    f = io.BytesIO()
    data = f.getvalue()
    response = requests.post(url,headers=HEADERS,params={"message": msg})

class transaction():
    def buy(numlistcurrency, amountbuy, shortlong):
        if test == False:
            try:
                client.create_order(symbol= currency[numlistcurrency] + exchange, side='BUY', type='MARKET', quantity=round(amountbuy, decimal[numlistcurrency])) 
                transaction.printData(numlistcurrency, "BUY" , amountbuy)
                if shortlong == "short": shortopenorder.writeandapend(numlistcurrency)
                elif shortlong == "long": openordertext.writeandapend(numlistcurrency)                
            except: transaction.printData(numlistcurrency, "BUY" , amountbuy), printError()
        else:
            try:
                print("test")
                client.create_test_order(symbol= currency[numlistcurrency] + exchange, side='BUY', type='MARKET', quantity=round(amountbuy, decimal[numlistcurrency])), transaction.printData(numlistcurrency, "BUY" , amountbuy)
            except: printError()
    def sell(numlistcurrency, amountsell, shortlong):
        if test == False:
            try:
                client.create_order(symbol= currency[numlistcurrency] + exchange, side='SELL', type='MARKET', quantity=round(amountsell, decimal[numlistcurrency]))
                transaction.printData(numlistcurrency, "SELL" , amountsell)                
                if shortlong == "short": shortopenorder.removelist(numlistcurrency)
                elif shortlong == "long": openordertext.removelist(numlistcurrency)
            except: transaction.printData(numlistcurrency, "SELL" , amountsell), printError()
        else:
            try:
                print("test")
                client.create_test_order(symbol= currency[numlistcurrency] + exchange, side='SELL', type='MARKET', quantity=round(amountsell, decimal[numlistcurrency])), transaction.printData(numlistcurrency, "SELL" , amountsell)
            except: printError()
    def printData(numlistcurrency, transactionSTR ,amount):
        printDataFormat = currency[numlistcurrency] ,datetimestrft() ,transactionSTR, "/price ", round(float(data(numlistcurrency).price), 3), "/amount ", round(amount, 5)
        print(printDataFormat)
        linesend(printDataFormat)

def terminalclear():
    import os
    os.system('cls' if os.name == 'nt' else 'clear')
    time.sleep(0.01)

class graphlong:
    def getGraph(coin):
        from numpy import average
        import yfinance as yf
        plt.gca().set_ylim([0, 3])
        ticker = yf.download((coin+'-USD'), dateback, datetime.datetime.now(), interval="1d")
        puellMultiple = []
        for x in range(int(setupNote('amountday'))+1, len(ticker['Close'])):
            past = []
            for y in range(x - int(setupNote('amountday')), x):
                past.append(ticker['Close'][y])
            puellMultiple.append(average(past) / ticker['Close'][x])
        plt.plot(puellMultiple, color = '#340B8C')
        plt.title(coin+'-USD')
        plt.gca().invert_yaxis()
        plt.show()
        print("")
    def getRECCOMENDATION(coinname):
        from numpy import average
        import yfinance as yf    
        ticker = yf.download((str(coinname)+'-USD'), dateback, datetime.datetime.now(), progress=False, interval="1d")
        plt.gca().set_ylim([0, 2.5])
        puellMultiple = []
        for x in range(int(setupNote('amountday'))+1, len(ticker['Close'])):
            past = []
            for y in range(x - int(setupNote('amountday')), x):
                past.append(ticker['Close'][y])
            puellMultiple.append(average(past) / ticker['Close'][x])
        buyvalue = average(puellMultiple) + (average(puellMultiple)* 38) /100
        sellvalue = average(puellMultiple) - (average(puellMultiple)* 34) /100
        if puellMultiple[len(puellMultiple)-1] > buyvalue: return {"COIN": coinname,"RECCOMMENDATION": "BUY", "CURRENT": puellMultiple[len(puellMultiple)-1], "AVERAGE": average(puellMultiple), "BUY_AT": buyvalue, "SELL_AT": sellvalue}
        elif puellMultiple[len(puellMultiple)-1] < sellvalue: return {"COIN": coinname, "RECCOMMENDATION":"SELL", "CURRENT": puellMultiple[len(puellMultiple)-1], "AVERAGE": average(puellMultiple), "BUY_AT": buyvalue, "SELL_AT": sellvalue}
        else: return {"COIN": coinname, "RECCOMMENDATION":"NUETRAL", "CURRENT": puellMultiple[len(puellMultiple)-1], "AVERAGE": average(puellMultiple), "BUY_AT": buyvalue, "SELL_AT": sellvalue}

class graphshort:
    def getGraph(coin):
        from numpy import average
        import yfinance as yf
        plt.gca().set_ylim([0, 3])
        ticker = yf.download((coin+'-USD'), datebackshort, datetime.datetime.now(), interval="1d")
        puellMultiple = []
        for x in range(int(setupNote('amountdayshort'))+1, len(ticker['Close'])):
            past = []
            for y in range(x - int(setupNote('amountdayshort')), x):
                past.append(ticker['Close'][y])
            puellMultiple.append(average(past) / ticker['Close'][x])
        plt.plot(puellMultiple, color = '#340B8C')
        plt.title(coin+'-USD')
        plt.gca().invert_yaxis()
        plt.show()
        print("")
    def getRECCOMENDATION(coinname):
        from numpy import average
        import yfinance as yf    
        ticker = yf.download((str(coinname)+'-USD'), datebackshort, datetime.datetime.now(), progress=False, interval="1d")
        plt.gca().set_ylim([0, 2.5])
        puellMultiple = []
        for x in range(int(setupNote('amountdayshort'))+1, len(ticker['Close'])):
            past = []
            for y in range(x - int(setupNote('amountdayshort')), x):
                past.append(ticker['Close'][y])
            puellMultiple.append(average(past) / ticker['Close'][x])        
        buyvalue = average(puellMultiple) + (average(puellMultiple)* float(14)) /100
        sellvalue = average(puellMultiple) - (average(puellMultiple)* float(15.2)) /100
        # buyvalue = average(puellMultiple) + (average(puellMultiple)* float(1)) /100
        # sellvalue = average(puellMultiple) - (average(puellMultiple)* float(0.3)) /100
        if puellMultiple[len(puellMultiple)-1] > buyvalue: return {"COIN": coinname,"RECCOMMENDATION": "BUY", "CURRENT": puellMultiple[len(puellMultiple)-1], "AVERAGE": average(puellMultiple), "BUY_AT": buyvalue, "SELL_AT": sellvalue}
        elif puellMultiple[len(puellMultiple)-1] < sellvalue: return {"COIN": coinname, "RECCOMMENDATION":"SELL", "CURRENT": puellMultiple[len(puellMultiple)-1], "AVERAGE": average(puellMultiple), "BUY_AT": buyvalue, "SELL_AT": sellvalue}
        else: return {"COIN": coinname, "RECCOMMENDATION":"NUETRAL", "CURRENT": puellMultiple[len(puellMultiple)-1], "AVERAGE": average(puellMultiple), "BUY_AT": buyvalue, "SELL_AT": sellvalue}

class longstart():
    def strat(limitopenorder):        
        df = []
        for x in range(len(currency)):
            df.append(graphlong.getRECCOMENDATION(currency[x]))
            if len(openordertext.openorder()) < int(limitopenorder) and all(t in openordertext.openorder() for t in [x]) == False:
                if graphlong.getRECCOMENDATION(currency[x])["RECCOMMENDATION"] == "BUY":
                    amountbuy = ((float(data(x).balanceUSDT)*((setupNote('percent')* (len(openordertext.openorder())+(len(shortopenorder.openorder()))+1))/100)) / float(data(x).price))
                    print(graphlong.getRECCOMENDATION(currency[x]))
                    transaction.buy(x, amountbuy, "long")
        for x in range (len(openordertext.openorder())):
                if graphlong.getRECCOMENDATION(currency[openordertext.openorder()[x]])["RECCOMMENDATION"] == "SELL":
                    print(graphlong.getRECCOMENDATION(currency[openordertext.openorder()[x]]))
                    minus = 0
                    for d in range(decimal[openordertext.openorder()[x]]): minus = minus / 10
                    amountsell = (float(data(openordertext.openorder()[x]).balanceCoin) - minus)
                    transaction.sell(openordertext.openorder()[x], amountsell, "long")
        terminalclear()
        print(datetimestrft(), ">>> OpenOrder:", len(openordertext.openorder()), "// LimitOpenOrder:", limitopenorder ,"\n")
        print(pandas.DataFrame(data=df), "\n")

class shortstrat:
    def strat(limitopenorder):
        df = []
        for x in range(len(currency)):
            df.append(graphshort.getRECCOMENDATION(currency[x]))
            if len(shortopenorder.openorder()) < int(limitopenorder) and all(t in shortopenorder.openorder() for t in [x]) == False:
                if graphshort.getRECCOMENDATION(currency[x])["RECCOMMENDATION"] == "BUY" and graphlong.getRECCOMENDATION(currency[x])["RECCOMMENDATION"] == "NUETRAL":
                    if (len(openordertext.openorder())+(len(shortopenorder.openorder()))) < 1:
                        amountbuy = ((float(data(x).balanceUSDT)*((setupNote('percent')* (len(openordertext.openorder())+(len(shortopenorder.openorder()))+1))/100)) / float(data(x).price))
                    else: amountbuy = ((float(data(x).balanceUSDT)*((setupNote('percent')* (len(openordertext.openorder())+(len(shortopenorder.openorder()))))/100)) / float(data(x).price))
                    print(graphshort.getRECCOMENDATION(currency[x]))
                    transaction.buy(x, amountbuy, "short")

        for x in range (len(shortopenorder.openorder())):
            if graphshort.getRECCOMENDATION(currency[shortopenorder.openorder()[x]])["RECCOMMENDATION"] == "SELL" and graphlong.getRECCOMENDATION(currency[x])["RECCOMMENDATION"] == "NUETRAL":
                print(graphshort.getRECCOMENDATION(currency[shortopenorder.openorder()[x]]))
                minus = 0
                for d in range(decimal[shortopenorder.openorder()[x]]): minus = minus / 10
                amountsell = (float(data(shortopenorder.openorder()[x]).balanceCoin) - minus)
                transaction.sell(shortopenorder.openorder()[x], amountsell, "short")
        print("\nshort\n")
        print(datetimestrft(), ">>> OpenOrder:", len(shortopenorder.openorder()), "// LimitOpenOrder:", limitopenorder ,"\n")
        print(pandas.DataFrame(data=df), "\n")

def typecommand(waitsec):
    import keyboard
    print("\n")
    for x in range(((int(waitsec))*10)+1):
        print("press Ctrl+Shift+A >>> Access command // SecCount ", x/10, "seconds", end="\r")
        if keyboard.is_pressed('ctrl+shift+a'):
            command = input("\ntradingAnalasis-04.py>>>   ")
            if command == 'plot':
                command = input("   short/long>>> ")
                if command == 'short':
                    command = input("   coin>>> ")
                    graphshort.getGraph(command)
                elif command == 'long':
                    command = input("   coin>>> ")
                    graphlong.getGraph(command)
            elif command == 'assets':
                from saveassets import saveassets
                saveassets.getdata(plot=True)
                
            else: pass
        else: pass
        time.sleep(0.1)

def runpennding():
    if '0455' <= datetime.datetime.now().strftime("%H%M") <= '0500':
        from saveassets import saveassets
        saveassets.insertdata()
        time.sleep(300)
        


# try: excelData()
# except:
#     printError()
#     pass
while True:
    try:
        longstart.strat(setupNote('limitOrder'))
        shortstrat.strat(setupNote('shortopenorder'))
        typecommand(30)
        runpennding()
              
        
    except:
        printError()
    
